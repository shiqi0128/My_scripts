from openpyxl import Workbook  # 导入一个openpyxl内的workbook方法
from openpyxl import load_workbook    # 打开已有文件
import datetime

# 1、创建一个Workbook对象
wb = Workbook("hello.xlxs")
# 激活worksheet
ws = wb.active
# 2、打开已有文件
wb2 = load_workbook("case.xlsx")
# 存储数据
# 数据可以直接分配到单元格中（可以输入公式）
ws['A1'] = 42
# 方式二: 可以附加行，从第一列开始附加（从最下方空白处，最左开始）（可以输入多行）
ws.append([1, 2, 3])
# 方式三:python类型会被自动转换
ws["A3"] = datetime.datetime.now().strftime("%Y-%m-%d")
# 3、创建表
# 方式一：插入到最后(default)
ws1 = wb.create_sheet("Mysheet")
# 方式二：插入到最开始的位置
ws2 = wb.create_sheet("Mysheet", 0)
# 4、选择表
# sheet 名称可以作为 key 进行索引
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
ws is ws3 is ws4
True


# 创建一个Sheet对象
mySheet = wb.create_sheet(index=0, title="Mysheet")


