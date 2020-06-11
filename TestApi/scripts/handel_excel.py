"""
-------------------------------------------------
  @Time : 2020/5/23 22:40 
  @Auth : 十七
  @File : handel_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""
from openpyxl import load_workbook


class HandelExcel:
    """封装一个excel的方法,方便后面使用的时候调用"""
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读数据
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        header_list = []
        testcase_list = []
        for row in range(1, ws.max_row+1):
            one_dict = {}
            for column in range(1, ws.max_column+1):
                one_cell = ws.cell(row, column)
                if row == 1:
                    header_list.append(one_cell.value)
                else:
                    key = header_list[column-1]
                    one_dict[key] = one_cell.value
            if row != 1:
                testcase_list.append(one_dict)

        return testcase_list

    def write_data(self, row, column, data):
        """
        写操作
        :param row:指定在某一行写
        :param column:指定在某一列写
        :param data:待写入的数据
        :return:
        """
        # 数据写入操作读写不能共用一个workbook和worksheet对象
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        # 第一种：
        one_cell = ws.cell(row, column)
        one_cell.value = data

        # 第二种：
        one_cell = ws.cell(row, column, value=data)
        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "testcase.xlsx"
    sheet_name = "login"
    do_excel = HandelExcel(excel_filename, sheet_name)
    # # print(do_excel.read_data())
    # do_excel.write_data(3, 6, 2)
