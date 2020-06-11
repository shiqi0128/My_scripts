# -*- coding: utf-8 -*-
"""
-------------------------------------------------
  @Time : 2020/4/29 21:38 
  @Auth : 可优
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: keyou100@qq.com
  @Company: 湖南省零檬信息技术有限公司
  @Copyright: 柠檬班
-------------------------------------------------
"""
# 1、xlxd、xlwt、openpyxl、pandas
# pip install -i https://pypi.douban.com/simple openpyxl==2.6.2
# a.workbook，往往用于创建一个新的excel文件，几乎不用
# from openpyxl import workbook
# b.load_workbook，往往对已存在的excel进行读写操作
from openpyxl import load_workbook

# c.openpyxl只能处理xlsx格式的excel文件
# d.只能使用办公软件来创建xlsx格式的excel文件，不能使用pycharm来创建
# e.excel对象 -> sheet表单对象 -> cell单元格对象 -> 行和列、值属性

# f.如果excel文件不存在，那么会FileNotFoundError
# res = load_workbook("testcase11.xlsx")

# 1、load_workbook函数，来加载excel文件
# a.第一个参数为excel文件名
# b.如果excel文件不存在，那么会FileNotFoundError
# c.返回Workbook对象，就相当于excel对象
wb = load_workbook("testcase.xlsx")

# 2、active默认读取第一个表单
# a.返回Worksheet，相当于sheet表单对象
ws = wb.active

# 3、读取指定单元格的数据
# a.cell方法，获取单元格，返回Cell对象
# b.Cell对象相当于cell单元格对象
one_cell = ws.cell(1, 1)
# c.可以使用Cell对象中的value来获取单元格的值

for row in range(1, ws.max_row):
    for column in range(1, ws.max_column):
        one_cell = ws.cell(row, column)
        print(f"单元格【{row}，{column}】中的值为{one_cell.value} 类型为{type(one_cell.value)}")

