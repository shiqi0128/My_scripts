# -*- coding: utf-8 -*-
"""
-------------------------------------------------
  @Time : 2020/4/29 21:38 
  @Auth : 可优
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: keyou100@qq.com
  @Company: 湖南省零檬信息技术有限公司
  @Copyright: 柠檬班
-------------------------------------------------
"""
# 0、xlxd、xlwt、openpyxl、pandas
from openpyxl import load_workbook

# 1、load_workbook函数，来加载excel文件
excel_filename = "testcase.xlsx"
wb = load_workbook(excel_filename)

# 2、active默认读取第一个表单
ws = wb['login']


# 3、写入指定单元格的数据
one_cell = ws.cell(4, 1)
one_cell.value = 3

# 4、保存excel
# a.Workbook对象.save("excel文件名")
# b.如果文件名与源文件不相同，那么会将源文件进行拷贝（赋值）
# wb.save("1.xlsx")
# c.PermissionError: [Errno 13] Permission denied: 'testcase.xlsx'
# 对exel文件修改之后，要保存，一定要将excel文件关闭
wb.save(excel_filename)
