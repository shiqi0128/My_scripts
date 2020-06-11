# -*- coding: utf-8 -*-
"""
-------------------------------------------------
  @Time : 2020/4/29 21:38 
  @Auth : 可优
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: keyou100@qq.com
  @Company: 湖南省零檬信息技术有限公司
  @Copyright: 柠檬班
-------------------------------------------------
"""
from openpyxl import load_workbook

# 1、load_workbook函数，来加载excel文件
# a.第一个参数为excel文件名
# b.如果excel文件不存在，那么会FileNotFoundError
# c.返回Workbook对象，就相当于excel对象
wb = load_workbook("testcase.xlsx")

# 2、active默认读取第一个表单
ws = wb['login']


# 3、读取指定单元格的数据
# 将数据读取出来，构造为嵌套字典的列表
# 存放用例字典
testcases_list = []
headers_list = []   # 存放表头信息
for row in range(1, ws.max_row+1):
    # 存放每一行的用例数据
    one_row_dict = {}
    for column in range(1, ws.max_column+1):
        # one_cell = ws.cell(row, column)
        one_cell_value = ws.cell(row, column).value
        if row == 1:
            headers_list.append(one_cell_value)
        else:
            key = headers_list[column - 1]
            one_row_dict[key] = one_cell_value
    # print(one_row_dict)
    # testcases_list.append(one_row_dict)

    if row != 1:
        testcases_list.append(one_row_dict)

# print(headers_list)
print(testcases_list)
