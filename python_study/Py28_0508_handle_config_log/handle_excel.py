# -*- coding: utf-8 -*-
"""
-------------------------------------------------
  @Time : 2020/5/6 20:44 
  @Auth : 可优
  @File : handle_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: keyou100@qq.com
  @Company: 湖南省零檬信息技术有限公司
  @Copyright: 柠檬班
-------------------------------------------------
"""
from openpyxl import load_workbook


# class Testcase:
#     pass


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """
        读数据
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        testcases_list = []
        headers_list = []  # 存放表头信息
        for row in range(1, ws.max_row + 1):
            # 存放每一行的用例数据
            one_row_dict = {}
            # one_testcase = Testcase()
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value
                # if one_cell_value is None:
                #     continue
                if row == 1:
                    headers_list.append(one_cell_value)
                else:
                    key = headers_list[column - 1]
                    one_row_dict[key] = one_cell_value
                    # setattr(one_testcase, str(key), one_cell_value)

            if row != 1:
                testcases_list.append(one_row_dict)
                # testcases_list.append(one_testcase)

        return testcases_list

    def write_data(self, row, column, data):
        """
        写操作
        :param row: 指定在某一行写
        :param column: 指定在某一列写
        :param data: 待写入的数据
        :return:
        """
        # 将数据写入到excel中，不能与读取操作公用一个Workbook对象
        # 如果使用同一个Workbook对象，只能将最后一次写入成功，会出现意想不到的结果
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        # 第一种写入方式：
        # one_cell = ws.cell(row, column)
        # one_cell.value = data

        # 第二种写入方式：
        ws.cell(row, column, value=data)

        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "testcase.xlsx"
    sheet_name = "login"
    do_excel = HandleExcel(excel_filename, sheet_name)
    print(do_excel.read_data())
    # do_excel.write_data(3, 6, 2)
