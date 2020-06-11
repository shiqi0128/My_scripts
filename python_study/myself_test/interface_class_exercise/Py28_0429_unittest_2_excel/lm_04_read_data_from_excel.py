"""
-------------------------------------------------
  @Time : 2020/5/22 14:00 
  @Auth : 十七
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""

from openpyxl import workbook    # workbook往往用于创建一个新的excel,几乎不用
from openpyxl import load_workbook   # load_workbook往往对已存在的excel进行读写操作

# 1、load_workbook来加载excel文件
# a、第一个参数为excel文件名
# b、如果excel文件不存在，那么会报错FileNotFoundError
# c、debug对象返回workbook对象，相当于excel
wb = load_workbook("testcase.xlsx")
pass
# 2、默认读取第一个表单，返回
# a、返回worksheet对象，相当于sheet对象
ws = wb.active

# 3、读取指定单元格的数据，先读取表单sheet对象后才能只能单元格对象
# a、cell方法，获取单元格，返回cell对象
# b、cell对象，相当于一个cell单元格对象
one_cell = ws.cell(1, 1)
# c、可以使用cell中的value来获取单元格的值
one_cell.value
pass

for row in range(1, ws.max_row):
    for column in range (1, ws.max_column):
        one_cell = ws.cell(row, column)
        print(f"单元格【{row, column}】中的值为{one_cell.value},类型为{type(one_cell.value)}")