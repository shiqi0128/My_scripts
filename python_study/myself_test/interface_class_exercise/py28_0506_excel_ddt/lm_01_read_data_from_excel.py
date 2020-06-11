"""
-------------------------------------------------
  @Time : 2020/5/22 14:00 
  @Auth : 十七
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""

from openpyxl import load_workbook


wb = load_workbook("testcase.xlsx")
# worksheet表单对象["sheet表单名"]可以直接获取到指定的表单
ws = wb["login"]
# ws_b = wb.active

for row in range(1, ws.max_row+1):
    for column in range(1, ws.max_column):
        one_cell = ws.cell(row, column)
        print(f"单元格【{row},{column}】的值为{one_cell.value}")
