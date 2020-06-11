"""
-------------------------------------------------
  @Time : 2020/5/22 21:39 
  @Auth : 十七
  @File : lm_02_write_data_into_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""
from openpyxl import load_workbook

wb = load_workbook("testcase.xlsx")
ws = wb["login"]
one_cell = ws.cell(5, 1)
one_cell.value = 4
# 4、保存excel
# a、workbook对象.save("excel文件名")
# b、如果文件名与源文件名不相同，那么会将源文件进行拷贝并赋值
# c、对源文件进行修改后一定要保存,打开excel一定要将excel先关闭执行脚本
# d、对excel文件进行读取的时候不需要关闭excel文件，进行写操作的时候一定要关闭excel
# wb.close()
wb.save("testcase.xlsx")
