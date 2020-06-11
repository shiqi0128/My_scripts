"""
-------------------------------------------------
  @Time : 2020/5/22 22:39 
  @Auth : 十七
  @File : py28_read_data_list_dict.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""
from openpyxl import load_workbook

wb = load_workbook("testcase.xlsx")


# sheet表单
ws = wb["login"]

# 单元格
# one_cell = ws.cell()
# 读取指定单元格的数据ll
testcase_list = []
header_list = []
for row in range(1, ws.max_row+1):
    # 存放每一行的用例数据
    one_row_dict = {}
    for column in range(1, ws.max_column+1):
        one_cell = ws.cell(row, column)
        if row == 1:
            header_list.append(one_cell.value)
        else:
            key = header_list[column-1]
            one_row_dict[key] = one_cell.value
    # print(one_row_dict)
    if row != 1:
        testcase_list.append(one_row_dict)
print(testcase_list)