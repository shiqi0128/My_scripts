"""
-------------------------------------------------
  @Time : 2020/5/22 15:28 
  @Auth : 十七
  @File : lm_04_read_data_from_excel.py
  @IDE  : PyCharm
  @Motto: ABC(Always Be Coding)
  @Email: 1941816343@qq.com
-------------------------------------------------
"""
from openpyxl import load_workbook

#   创建一个excel对象
wb = load_workbook("testcase.xlsx")
pass
# 读取表单数据
ws = wb.active

# 读取表单数据后才能读取单元格的数据
one_cell = ws.cell(1, 1)

# ws.max_row，是读取表单数据后取最大行
for row in range(1, ws.max_row+1):
    # ws.max_column,先读取表单数据后的最大列
    for column in range(1, ws.max_column+1):
        one_cell = ws.cell(row, column)
        print(f"单元格[{row, column}]的值为{one_cell.value}, 类型为{type(one_cell.value)}")